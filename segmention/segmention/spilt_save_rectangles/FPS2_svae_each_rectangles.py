'''
Descripttion: 
version: 
Author: lyh
Date: 2023-03-15 14:24:52
LastEditors: smile
LastEditTime: 2023-06-01 07:37:34
'''
import cv2
import numpy as np
import openpyxl

workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = 'Positions'
img = gray= cv2.imread(r'VIS1.jpg')
raw_IRimg = cv2.imread(r'IR1.jpg')   
raw_VISimg = cv2.imread(r'VIS1.jpg')
print(img.shape,gray.shape)             
sift = cv2.SIFT_create()
kp, des = sift.detectAndCompute(gray, None)
rect_width = 100
rect_height = 100
rectangles = []
for i in range(len(kp)):
    x, y = np.int32(kp[i].pt)
    left = np.int32(x - rect_width / 2)
    top = np.int32(y - rect_height / 2)
    right = np.int32(x + rect_width / 2)
    bottom = np.int32(y + rect_height / 2)
    is_intersect = False
    for rect in rectangles:
        if (left < rect[2] and right > rect[0] and top < rect[3] and bottom > rect[1]):
            is_intersect = True
            break
    if not is_intersect:
        rectangles.append([left, top, right, bottom])
worksheet.cell(row=1, column=1).value = 'rect[0]'
worksheet.cell(row=1, column=2).value = 'rect[1]'
worksheet.cell(row=1, column=3).value = 'rect[2]'
worksheet.cell(row=1, column=4).value = 'rect[3]'
worksheet.cell(row=1, column=5).value = 'number'
raw_VISimg_hole = raw_VISimg
raw_IRimg_hole = raw_IRimg
for i, rect in enumerate(rectangles):
    roi_IR = raw_IRimg[rect[1]:rect[3], rect[0]:rect[2]]
    roi_VIS = raw_VISimg[rect[1]:rect[3], rect[0]:rect[2]]
    #print(roi_VIS.shape)
    try:
        cv2.imwrite('spilt_img/spilt_fusion/IR_spilt/IR_spilt_subrect/ir_rectangle_{}.png'.format(i), roi_IR)
        cv2.imwrite('spilt_img/spilt_fusion/IR_spilt/IR_spilt_subrect/vis_rectangle_{}.png'.format(i), roi_VIS)
        print("第{}".format(i))
        # 将位置写入Excel
        worksheet.cell(row=i+1, column=1).value = rect[0]
        worksheet.cell(row=i+1, column=2).value = rect[1]
        worksheet.cell(row=i+1, column=3).value = rect[2]
        worksheet.cell(row=i+1, column=4).value = rect[3]
        worksheet.cell(row=i+1, column=5).value = i
        raw_VISimg_hole[rect[1]:rect[3], rect[0]:rect[2]] = 0
        raw_IRimg_hole[rect[1]:rect[3], rect[0]:rect[2]] = 0
    except cv2.error:
        print('Error saving image: ir_rectangle_{}.png'.format(i))
        continue
cv2.imwrite('spilt_img/spilt_fusion/VIS_spilt/Hole_VISimg_linshi.png',raw_VISimg_hole)
cv2.imwrite('spilt_img/spilt_fusion/VIS_spilt/Hole_IRimg_linshi.png',raw_IRimg_hole)
for row in worksheet:
    if all(cell.value is None for cell in row):
        worksheet.delete_rows(row[0].row)
workbook.save('spilt_img/spilt_fusion/spilt_save_rectangles/positions——linshi.xlsx')
